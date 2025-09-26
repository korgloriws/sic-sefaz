import io
import unicodedata
import re
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ----------------------------
# Helpers: normalization
# ----------------------------
def _normalize_text(value: str) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
  
    text = "".join(ch if ch.isalnum() else "" for ch in text)
    return text


def _build_normalized_columns_map(df: pd.DataFrame) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for col in df.columns:
        mapping[_normalize_text(col)] = col
    return mapping


def _rename_columns_using_synonyms(
    df: pd.DataFrame,
    expected_to_synonyms: Dict[str, Set[str]],
) -> Tuple[pd.DataFrame, Dict[str, str]]:
    normalized_map = _build_normalized_columns_map(df)
    rename_map: Dict[str, str] = {}
    found: Dict[str, str] = {}

    for expected_key, synonyms in expected_to_synonyms.items():
        target_original_name: Optional[str] = None
        for syn in synonyms:
            syn_norm = _normalize_text(syn)
            if syn_norm in normalized_map:
                target_original_name = normalized_map[syn_norm]
                break
        if target_original_name:
            rename_map[target_original_name] = expected_key
            found[expected_key] = target_original_name

    df2 = df.rename(columns=rename_map)
    return df2, found


def _ensure_required_columns(
    df: pd.DataFrame, required: List[str], context_label: str
) -> None:
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            f"Colunas ausentes para {context_label}: {', '.join(missing)}. "
            "Verifique os nomes das colunas do arquivo."
        )


# ----------------------------
# Helpers: numeric coercion
# ----------------------------
def _coerce_numeric(series: pd.Series) -> pd.Series:
    import re
    def parse_value(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return pd.NA
        if isinstance(x, (int, float)):

            return float(x)
        s = str(x).strip()
        if s == "":
            return pd.NA
       
        s = s.replace("\u00A0", "").replace(" ", "")
   
        negative = False
        if s.startswith("(") and s.endswith(")"):
            negative = True
            s = s[1:-1]
       
        s = re.sub(r"[^0-9,\.-]", "", s)

        if "," in s and "." in s:
      
            if s.rfind(",") > s.rfind("."):
             
                s2 = s.replace(".", "").replace(",", ".")
            else:
             
                s2 = s.replace(",", "")
        elif "," in s:
          
            s2 = s.replace(".", "").replace(",", ".")
        elif "." in s:
          
            s2 = s
        else:
            
            if len(s) <= 2:
                s2 = "0." + s.zfill(2)
            else:
                s2 = s[:-2] + "." + s[-2:]

        try:
            val = float(s2)
            if negative:
                val = -val
            return val
        except Exception:
            return pd.NA

    return series.apply(parse_value).astype(float)


# ----------------------------
# Core logic
# ----------------------------
def _infer_tipo_conta_column(df: pd.DataFrame) -> Optional[str]:
    # Try by header keywords first
    name_candidates = [
        "tipoconta",
        "tipo_conta",
        "tipocont",
        "tipodaconta",
        "tpconta",
        "tp_conta",
        "tpcta",
        "tipocta",
        "tipoc",
    ]
    normalized_map = _build_normalized_columns_map(df)
    for norm_name, original in normalized_map.items():
        if any(k in norm_name for k in name_candidates):
            return original


    for col in df.columns:
        s = df[col].dropna().astype(str).str.strip()
        if s.empty:
            continue
        digits_mask = s.str.match(r"^[0-9]$")
        if digits_mask.mean() >= 0.8:  
          
            if (s == "5").mean() >= 0.05:
                return col
   
    for col in df.columns:
        s = df[col].dropna().astype(str).str.strip()
        if s.empty:
            continue
        if s.str.match(r"^[0-9]$").mean() >= 0.8:
            return col
    return None


def _infer_atributo_column(df: pd.DataFrame) -> Optional[str]:

    name_candidates = [
        "atributo",
        "atr",
        "atrib",
        "attr",
        "indicador",
        "ind",
        "flag",
    ]
    normalized_map = _build_normalized_columns_map(df)
    for norm_name, original in normalized_map.items():
        if any(k in norm_name for k in name_candidates):
            return original

   
    for col in df.columns:
        s = df[col].dropna().astype(str).str.strip().str.upper()
        if s.empty:
            continue
        valid = s.isin(["F", "T"])
        if valid.mean() >= 0.8:
            return col
    return None


def _standardize_990(df_990: pd.DataFrame) -> pd.DataFrame:
    synonyms: Dict[str, Set[str]] = {
        "cod_contabil": {
            "cod_contabil",
            "codigo_contabil",
            "c\u00f3digo cont\u00e1bil",
            "conta_contabil",
            "contacontabil",
            "codcontabil",
        },
        "tipo_conta": {
            "tipo_conta",
            "tipo",
            "tipo de conta",
            "tipoconta",
            "tp_conta",
            "tpconta",
            "tpcta",
            "tipo cta",
            "tipo da conta",
        },
        "atributo": {
            "atributo",
            "attr",
            "atr",
            "atrib",
            "indicador",
            "ind",
            "flag",
        },
    }
    df2, found = _rename_columns_using_synonyms(df_990, synonyms)


    required_cols = ["cod_contabil", "tipo_conta", "atributo"]
    present = set(df2.columns)
    missing_set = [c for c in required_cols if c not in present]
    if missing_set:
       
        if "tipo_conta" in missing_set:
            inferred = _infer_tipo_conta_column(df_990)
            if inferred and inferred in df2.columns:
                df2 = df2.rename(columns={inferred: "tipo_conta"})
            elif inferred and inferred not in df2.columns and inferred in df_990.columns:
               
                df2["tipo_conta"] = df_990[inferred]
        if "atributo" in missing_set:
            inferred = _infer_atributo_column(df_990)
            if inferred and inferred in df2.columns:
                df2 = df2.rename(columns={inferred: "atributo"})
            elif inferred and inferred not in df2.columns and inferred in df_990.columns:
                df2["atributo"] = df_990[inferred]

    _ensure_required_columns(df2, required_cols, "990")

 
    df2["cod_contabil"] = df2["cod_contabil"].astype(str).str.strip()
    df2["tipo_conta"] = df2["tipo_conta"].astype(str).str.strip()
    df2["atributo"] = df2["atributo"].astype(str).str.strip().str.upper()
    return df2


def _standardize_balancete(df_bal: pd.DataFrame) -> pd.DataFrame:
    synonyms: Dict[str, Set[str]] = {
        "cod_contabil": {
            "cod_contabil",
            "codigo_contabil",
            "c\u00f3digo cont\u00e1bil",
            "conta_contabil",
            "contacontabil",
            "codcontabil",
        },
        "debito_atual": {
            "debito_atual",
            "debito",
            "debito atual",
            "vl_debito_atual",
            "vlr_debito_atual",
            "saldo_debito_atual",
            "deb_atual",
        },
        "credito_atual": {
            "credito_atual",
            "cr\u00e9dito_atual",
            "credito",
            "credito atual",
            "vl_credito_atual",
            "vlr_credito_atual",
            "saldo_credito_atual",
            "cred_atual",
        },
    }
    df2, found = _rename_columns_using_synonyms(df_bal, synonyms)
    _ensure_required_columns(
        df2, ["cod_contabil", "debito_atual", "credito_atual"], "balancete"
    )

    df2["cod_contabil"] = df2["cod_contabil"].astype(str).str.strip()
    
    return df2


def _filter_990_tipo5_atributoF(df_990_std: pd.DataFrame) -> pd.DataFrame:
    mask = (df_990_std["tipo_conta"].astype(str).str.strip() == "5") & (
        df_990_std["atributo"].astype(str).str.strip().str.upper() == "F"
    )
    return df_990_std.loc[mask].copy()


def _filter_990_tipo5(df_990_std: pd.DataFrame) -> pd.DataFrame:
    mask = df_990_std["tipo_conta"].astype(str).str.strip() == "5"
    return df_990_std.loc[mask].copy()


def _sum_for_prefix(
    df_bal_std: pd.DataFrame,
    allowed_990_codes: Set[str],
    prefix: str,
    value_column: str,
    match_mode: str = "equals",
) -> float:
    def _norm_code(x: str) -> str:
        return re.sub(r"\D", "", str(x or "").strip())

    bal_codes_raw = df_bal_std["cod_contabil"].astype(str).str.strip()
    bal_codes = bal_codes_raw.apply(_norm_code)
    prefix_norm = _norm_code(prefix)
    mask_prefix = bal_codes.str.startswith(prefix_norm)

    if not allowed_990_codes:
        return 0.0

    allowed_norm = {_norm_code(c) for c in allowed_990_codes}

    allowed_trim = {c.rstrip("0") for c in allowed_norm if c}
    allowed_all = {c for c in allowed_norm if c} | {c for c in allowed_trim if c}

    if match_mode == "equals":
        rows_mask = mask_prefix & bal_codes.isin(allowed_all)
    else:
        allowed_list = list(allowed_all)
        allowed_prefix_mask = bal_codes.apply(
            lambda c: any(c.startswith(ac) for ac in allowed_list if ac)
        )
        rows_mask = mask_prefix & allowed_prefix_mask

    if not rows_mask.any():
        return 0.0

    values = _coerce_numeric(df_bal_std.loc[rows_mask, value_column])
    total = float(values.sum(skipna=True))
    return total


def _sum_prefix_only(
    df_bal_std: pd.DataFrame,
    prefix: str,
    value_column: str,
) -> float:
    import re
    def _norm_code(x: str) -> str:
        return re.sub(r"\D", "", str(x or "").strip())
    bal_codes_raw = df_bal_std["cod_contabil"].astype(str).str.strip()
    bal_codes = bal_codes_raw.apply(_norm_code)
    prefix_norm = _norm_code(prefix)
    rows_mask = bal_codes.str.startswith(prefix_norm)
    if not rows_mask.any():
        return 0.0
    values = _coerce_numeric(df_bal_std.loc[rows_mask, value_column])
    return float(values.sum(skipna=True))


def _sum_diff_for_prefix(
    df_bal_std: pd.DataFrame,
    allowed_990_codes: Set[str],
    prefix: str,
    match_mode: str = "prefix",
) -> float:
    import re
    def _norm_code(x: str) -> str:
        return re.sub(r"\D", "", str(x or "").strip())

    bal_codes_raw = df_bal_std["cod_contabil"].astype(str).str.strip()
    bal_codes = bal_codes_raw.apply(_norm_code)
    prefix_norm = _norm_code(prefix)
    mask_prefix = bal_codes.str.startswith(prefix_norm)

    if not allowed_990_codes:
        rows_mask = mask_prefix
    else:
        allowed_norm = {_norm_code(c) for c in allowed_990_codes}
        allowed_trim = {c.rstrip("0") for c in allowed_norm if c}
        allowed_all = {c for c in allowed_norm if c} | {c for c in allowed_trim if c}
        if match_mode == "equals":
            rows_mask = mask_prefix & bal_codes.isin(allowed_all)
        else:
            allowed_list = list(allowed_all)
            allowed_prefix_mask = bal_codes.apply(
                lambda c: any(c.startswith(ac) for ac in allowed_list if ac)
            )
            rows_mask = mask_prefix & allowed_prefix_mask

    if not rows_mask.any():
        return 0.0

    cred = _coerce_numeric(df_bal_std.loc[rows_mask, "credito_atual"]).sum(skipna=True)
    deb = _coerce_numeric(df_bal_std.loc[rows_mask, "debito_atual"]).sum(skipna=True)
    return float((cred if pd.notna(cred) else 0.0) - (deb if pd.notna(deb) else 0.0))


def _sum_diff_prefix_only(
    df_bal_std: pd.DataFrame,
    prefix: str,
) -> float:
    import re
    def _norm_code(x: str) -> str:
        return re.sub(r"\D", "", str(x or "").strip())
    bal_codes_raw = df_bal_std["cod_contabil"].astype(str).str.strip()
    bal_codes = bal_codes_raw.apply(_norm_code)
    prefix_norm = _norm_code(prefix)
    rows_mask = bal_codes.str.startswith(prefix_norm)
    if not rows_mask.any():
        return 0.0
    cred = _coerce_numeric(df_bal_std.loc[rows_mask, "credito_atual"]).sum(skipna=True)
    deb = _coerce_numeric(df_bal_std.loc[rows_mask, "debito_atual"]).sum(skipna=True)
    return float((cred if pd.notna(cred) else 0.0) - (deb if pd.notna(deb) else 0.0))


def compute_campos(df_990: pd.DataFrame, df_balancete: pd.DataFrame) -> List[float]:
    df990_std = _standardize_990(df_990)
    dfbal_std = _standardize_balancete(df_balancete)


    df990_f = _filter_990_tipo5_atributoF(df990_std)
    allowed_codes_f = set(df990_f["cod_contabil"].dropna().astype(str).str.strip())


    df990_t5 = _filter_990_tipo5(df990_std)
    allowed_codes_t5 = set(df990_t5["cod_contabil"].dropna().astype(str).str.strip())


    campo1 = _sum_for_prefix(dfbal_std, allowed_codes_f, "1", "debito_atual", match_mode="equals")
   
    campo2 = _sum_for_prefix(dfbal_std, allowed_codes_f, "2", "credito_atual", match_mode="equals")
 
    campo3 = _sum_diff_for_prefix(dfbal_std, allowed_codes_t5, "6221301", match_mode="prefix")
    if campo3 == 0.0:
        fallback3 = _sum_diff_prefix_only(dfbal_std, "6221301")
        if fallback3 != 0.0:
            campo3 = fallback3
 
    campo4 = _sum_diff_for_prefix(dfbal_std, allowed_codes_t5, "6311", match_mode="prefix")
    if campo4 == 0.0:
        fallback4 = _sum_diff_prefix_only(dfbal_std, "6311")
        if fallback4 != 0.0:
            campo4 = fallback4
 
    campo5 = _sum_diff_for_prefix(dfbal_std, allowed_codes_t5, "8211101", match_mode="prefix")
    if campo5 == 0.0:
        fallback5 = _sum_diff_prefix_only(dfbal_std, "8211101")
        if fallback5 != 0.0:
            campo5 = fallback5

    return [campo1, campo2, campo3, campo4, campo5]


def _aggregate_sum_by_code(
    df_bal_std: pd.DataFrame,
    rows_mask: pd.Series,
    value_column: str,
) -> pd.DataFrame:
    subset = df_bal_std.loc[rows_mask, ["cod_contabil", value_column]].copy()
    subset[value_column] = _coerce_numeric(subset[value_column])
    grouped = (
        subset.groupby("cod_contabil", as_index=False)[value_column]
        .sum()
        .rename(columns={value_column: "valor"})
    )
    return grouped


def _aggregate_diff_by_code(
    df_bal_std: pd.DataFrame,
    rows_mask: pd.Series,
) -> pd.DataFrame:
    subset = df_bal_std.loc[rows_mask, ["cod_contabil", "credito_atual", "debito_atual"]].copy()
    subset["credito_atual"] = _coerce_numeric(subset["credito_atual"])
    subset["debito_atual"] = _coerce_numeric(subset["debito_atual"])
    grouped = (
        subset.groupby("cod_contabil", as_index=False)
        .agg({"credito_atual": "sum", "debito_atual": "sum"})
    )
    grouped["valor"] = grouped["credito_atual"].fillna(0.0) - grouped["debito_atual"].fillna(0.0)
    grouped = grouped[["cod_contabil", "valor"]]
    return grouped


def compute_campos_with_details(
    df_990: pd.DataFrame, df_balancete: pd.DataFrame
) -> Tuple[List[float], pd.DataFrame]:
    df990_std = _standardize_990(df_990)
    dfbal_std = _standardize_balancete(df_balancete)

    df990_f = _filter_990_tipo5_atributoF(df990_std)
    allowed_codes_f = set(df990_f["cod_contabil"].dropna().astype(str).str.strip())

    df990_t5 = _filter_990_tipo5(df990_std)
    allowed_codes_t5 = set(df990_t5["cod_contabil"].dropna().astype(str).str.strip())


    def _norm_code(x: str) -> str:
        return re.sub(r"\D", "", str(x or "").strip())

    bal_codes_raw = dfbal_std["cod_contabil"].astype(str).str.strip()
    bal_codes = bal_codes_raw.apply(_norm_code)


    def _allowed_all(codes: Set[str]) -> Set[str]:
        norm = {_norm_code(c) for c in codes}
        trim = {c.rstrip("0") for c in norm if c}
        return {c for c in norm if c} | {c for c in trim if c}

    allowed_f_all = _allowed_all(allowed_codes_f)
    allowed_t5_all = _allowed_all(allowed_codes_t5)

    detalhes_rows: List[Dict[str, object]] = []


    prefix = _norm_code("1")
    mask_prefix = bal_codes.str.startswith(prefix)
    rows_mask_1 = mask_prefix & bal_codes.isin(allowed_f_all)
    det1 = _aggregate_sum_by_code(dfbal_std, rows_mask_1, "debito_atual")
    det1_total = float(det1["valor"].sum()) if not det1.empty else 0.0
    det1["campo"] = 1

  
    prefix = _norm_code("2")
    mask_prefix = bal_codes.str.startswith(prefix)
    rows_mask_2 = mask_prefix & bal_codes.isin(allowed_f_all)
    det2 = _aggregate_sum_by_code(dfbal_std, rows_mask_2, "credito_atual")
    det2_total = float(det2["valor"].sum()) if not det2.empty else 0.0
    det2["campo"] = 2

   
    prefix = _norm_code("6221301")
    mask_prefix = bal_codes.str.startswith(prefix)
    if allowed_t5_all:
        allowed_list = list(allowed_t5_all)
        allowed_prefix_mask = bal_codes.apply(lambda c: any(c.startswith(ac) for ac in allowed_list if ac))
        rows_mask_3 = mask_prefix & allowed_prefix_mask
    else:
        rows_mask_3 = mask_prefix
    det3 = _aggregate_diff_by_code(dfbal_std, rows_mask_3)
    det3_total = float(det3["valor"].sum()) if not det3.empty else 0.0
    if det3_total == 0.0:
    
        rows_mask_3_fb = mask_prefix
        det3 = _aggregate_diff_by_code(dfbal_std, rows_mask_3_fb)
        det3_total = float(det3["valor"].sum()) if not det3.empty else 0.0
    det3["campo"] = 3

 
    prefix = _norm_code("6311")
    mask_prefix = bal_codes.str.startswith(prefix)
    if allowed_t5_all:
        allowed_list = list(allowed_t5_all)
        allowed_prefix_mask = bal_codes.apply(lambda c: any(c.startswith(ac) for ac in allowed_list if ac))
        rows_mask_4 = mask_prefix & allowed_prefix_mask
    else:
        rows_mask_4 = mask_prefix
    det4 = _aggregate_diff_by_code(dfbal_std, rows_mask_4)
    det4_total = float(det4["valor"].sum()) if not det4.empty else 0.0
    if det4_total == 0.0:
        rows_mask_4_fb = mask_prefix
        det4 = _aggregate_diff_by_code(dfbal_std, rows_mask_4_fb)
        det4_total = float(det4["valor"].sum()) if not det4.empty else 0.0
    det4["campo"] = 4


    prefix = _norm_code("8211101")
    mask_prefix = bal_codes.str.startswith(prefix)
    if allowed_t5_all:
        allowed_list = list(allowed_t5_all)
        allowed_prefix_mask = bal_codes.apply(lambda c: any(c.startswith(ac) for ac in allowed_list if ac))
        rows_mask_5 = mask_prefix & allowed_prefix_mask
    else:
        rows_mask_5 = mask_prefix
    det5 = _aggregate_diff_by_code(dfbal_std, rows_mask_5)
    det5_total = float(det5["valor"].sum()) if not det5.empty else 0.0
    if det5_total == 0.0:
        rows_mask_5_fb = mask_prefix
        det5 = _aggregate_diff_by_code(dfbal_std, rows_mask_5_fb)
        det5_total = float(det5["valor"].sum()) if not det5.empty else 0.0
    det5["campo"] = 5

    valores = [det1_total, det2_total, det3_total, det4_total, det5_total]
    detalhes_df = pd.concat([det1, det2, det3, det4, det5], ignore_index=True)

    detalhes_df = detalhes_df[["campo", "cod_contabil", "valor"]]

    try:
        detalhes_df = detalhes_df.sort_values(["campo", "cod_contabil"]).reset_index(drop=True)
    except Exception:
        pass
    return valores, detalhes_df


def _prepare_template_bytes(template_file, valores: List[float], detalhes: Optional[pd.DataFrame] = None) -> bytes:
  
    try:
        template_file.seek(0)
        data = template_file.read()
    except Exception:
        data = template_file.getvalue() if hasattr(template_file, "getvalue") else None
    if data is None:
        raise ValueError("Falha ao ler o arquivo de confer\u00eancias (template).")

    in_buf = io.BytesIO(data)
    wb = load_workbook(in_buf)
    ws = wb.active

   
    targets = ["B2", "B3", "B4", "B5", "B7"]
    for cell_ref, value in zip(targets, valores):
        ws[cell_ref] = value

    
    if detalhes is not None and not detalhes.empty:
        ws2 = wb.create_sheet(title="Detalhes")
        headers = ["Campo", "Cod_Contabil", "Valor"]
        ws2.append(headers)
        for _, row in detalhes.iterrows():
            ws2.append([int(row["campo"]), str(row["cod_contabil"]), float(row["valor"])])

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------
def main() -> None:
    st.title("Comparador recursos disponiveis")

    st.markdown(
        "Envie os 3 arquivos XLSX: 1) 990 (Plano de Contas), 2) Balancete, 3) Confer\u00eancias DDR (template a ser preenchido)."
    )

    col1, col2 = st.columns(2)
    with col1:
        file_990 = st.file_uploader(
            "Arquivo 990 (plano de contas)", type=["xlsx"], key="file_990"
        )
        file_bal = st.file_uploader(
            "Arquivo Balancete", type=["xlsx"], key="file_bal"
        )
    with col2:
        file_conf = st.file_uploader(
            "Arquivo Confer\u00eancias DDR (template)", type=["xlsx"], key="file_conf"
        )

    if st.button("Processar e Preencher", type="primary"):
        if not file_990 or not file_bal or not file_conf:
            st.error("Por favor, envie os 3 arquivos para continuar.")
            return

        try:

            df_990 = pd.read_excel(file_990, dtype=object)
            file_990.seek(0)
            df_bal = pd.read_excel(file_bal, dtype=object)
            file_bal.seek(0)

            valores, detalhes_df = compute_campos_with_details(df_990, df_bal)


            st.subheader("Resultados Computados")
            labels = [
                "Campo 1 (prefixo 1, d\u00e9bito atual)",
                "Campo 2 (prefixo 2, cr\u00e9dito atual)",
                "Campo 3 (prefixo 6221301, cr\u00e9dito atual)",
                "Campo 4 (prefixo 6311, cr\u00e9dito atual)",
                "Campo 5 (prefixo 8211101, cr\u00e9dito atual)",
            ]
            st.dataframe(
                pd.DataFrame({"Campo": labels, "Valor": valores}),
                use_container_width=True,
            )


            filled_bytes = _prepare_template_bytes(file_conf, valores, detalhes_df)

            st.download_button(
                label="Baixar Confer\u00eancias DDR preenchido",
                data=filled_bytes,
                file_name="Conferencias_DDR_preenchido.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )

            st.success("Arquivo preenchido gerado com sucesso.")
        except Exception as exc:
            st.exception(exc)


if __name__ == "__main__":
    main()


