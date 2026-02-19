

import io
import csv
import re
from collections import defaultdict

import streamlit as st
from openpyxl import Workbook


DELIMITER = ";"
CTB_REGISTRO_FILTRO = "20"
CTB_COL_REGISTRO = 0
CTB_COL_COD_REDUZIDO = 2
CTB_COL_FONTE = 3
CTB_COL_VALOR = 6

POSICAO_COL_COD_REDUZIDO = "cod_reduz_banco"
POSICAO_COL_FONTE = "fonte"
POSICAO_COL_VALOR = "saldo_real"


def extrair_nucleo_fonte(fonte: str) -> str:

    if not fonte or not isinstance(fonte, str):
        return ""
    dig = re.sub(r"\D", "", fonte)
    if len(dig) == 7:
        return dig[:4]
    if len(dig) == 8:
        return dig[1:5]
    return dig


def parse_valor_br(val: str) -> float:
   
    if not val or not isinstance(val, str):
        return 0.0
    s = val.strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def carregar_ctb_agregado(content: bytes, encoding: str = "utf-8") -> dict:

    agregado = defaultdict(float)
    try:
        text = content.decode(encoding)
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text), delimiter=DELIMITER)
    for row in reader:
        if len(row) <= max(CTB_COL_REGISTRO, CTB_COL_COD_REDUZIDO, CTB_COL_FONTE, CTB_COL_VALOR):
            continue
        if row[CTB_COL_REGISTRO].strip() != CTB_REGISTRO_FILTRO:
            continue
        cod = row[CTB_COL_COD_REDUZIDO].strip()
        fonte = row[CTB_COL_FONTE].strip()
        nucleo = extrair_nucleo_fonte(fonte)
        valor = parse_valor_br(row[CTB_COL_VALOR])
        agregado[(cod, nucleo)] += valor
    return dict(agregado)


def carregar_posicao_agregado(content: bytes, encoding: str = "utf-8") -> dict:

    agregado = defaultdict(float)
    try:
        text = content.decode(encoding)
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text), delimiter=DELIMITER)
    if POSICAO_COL_COD_REDUZIDO not in (reader.fieldnames or []):
        return {}
    for row in reader:
        cod = (row.get(POSICAO_COL_COD_REDUZIDO) or "").strip()
        fonte = (row.get(POSICAO_COL_FONTE) or "").strip()
        nucleo = extrair_nucleo_fonte(fonte)
        valor = parse_valor_br(row.get(POSICAO_COL_VALOR) or "0")
        agregado[(cod, nucleo)] += valor
    return dict(agregado)


def comparar(ctb: dict, posicao: dict):

    chaves = set(ctb) | set(posicao)
    resultado = []
    for (cod, nucleo) in sorted(chaves):
        v_ctb = ctb.get((cod, nucleo), 0.0)
        v_pos = posicao.get((cod, nucleo), 0.0)
        diff = v_ctb - v_pos
        resultado.append((cod, nucleo, v_ctb, v_pos, diff))
    return resultado


def somar_por_fonte(ctb: dict, posicao: dict) -> list:
    """Agrega totais por nucleo (fonte), somando todos os cod_reduzido de cada fonte."""
    total_ctb = defaultdict(float)
    total_pos = defaultdict(float)
    for (cod, nucleo) in ctb:
        total_ctb[nucleo] += ctb[(cod, nucleo)]
    for (cod, nucleo) in posicao:
        total_pos[nucleo] += posicao[(cod, nucleo)]
    nucleos = sorted(set(total_ctb) | set(total_pos))
    return [
        (nucleo, total_ctb[nucleo], total_pos[nucleo], total_ctb[nucleo] - total_pos[nucleo])
        for nucleo in nucleos
    ]


def main():


    st.title(" CTB x Posicao de Bancos por Fonte")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Arquivo CTB")
        st.caption("CSV sem cabecalho, separador ponto e virgula. Usado: registro 20, cod reduzido (C), fonte (D), valor (G).")
        file_ctb = st.file_uploader("Upload CTB", type=["csv"], key="ctb")
    with col2:
        st.subheader("Arquivo Posicao de Bancos")
        st.caption("CSV com cabecalho. Colunas: cod_reduz_banco, fonte, saldo_real.")
        file_posicao = st.file_uploader("Upload Posicao de Bancos", type=["csv"], key="posicao")

    if not file_ctb or not file_posicao:
        st.info("Envie os dois arquivos CSV para gerar a comparacao.")
        return

    ctb_bytes = file_ctb.read()
    posicao_bytes = file_posicao.read()

    ctb_agr = carregar_ctb_agregado(ctb_bytes)
    pos_agr = carregar_posicao_agregado(posicao_bytes)
    linhas = comparar(ctb_agr, pos_agr)
    soma_fonte = somar_por_fonte(ctb_agr, pos_agr)
    total_ctb = sum(ctb_agr.values())
    total_posicao = sum(pos_agr.values())
    diferenca_total = total_ctb - total_posicao

    st.subheader("Resultado da comparacao")
    st.caption("Agregado por (codigo reduzido + nucleo da fonte). Valores somados. Diferenca = CTB - Posicao.")

    if not linhas:
        st.warning("Nenhum registro encontrado apos filtros.")
        return


    diff_ok = sum(1 for r in linhas if abs(r[4]) < 1e-6)
    diff_div = len(linhas) - diff_ok

    # --- Grid: métricas da comparação em uma linha ---
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Pares (cod + fonte) no total", len(linhas))
    with m2:
        st.metric("Conferidos (diferenca zero)", diff_ok)
    with m3:
        st.metric("Divergentes", diff_div)

    # --- Tabela detalhada com scroll (altura fixa) ---
    df_comparacao = {
        "Cod. Reduzido": [r[0] for r in linhas],
        "Nucleo Fonte": [r[1] for r in linhas],
        "Valor CTB": [round(r[2], 2) for r in linhas],
        "Valor Posicao": [round(r[3], 2) for r in linhas],
        "Diferenca": [round(r[4], 2) for r in linhas],
    }
    st.dataframe(df_comparacao, height=320, use_container_width=True)

    # --- Grid: Somatório por fonte (tabela com scroll) | Soma total (métricas) ---
    col_fonte, col_total = st.columns([1, 1])
    with col_fonte:
        st.subheader("Somatório por fonte")
        st.caption("Total por nucleo (fonte).")
        df_fonte = {
            "Nucleo (Fonte)": [r[0] for r in soma_fonte],
            "Total CTB": [round(r[1], 2) for r in soma_fonte],
            "Total Posicao": [round(r[2], 2) for r in soma_fonte],
            "Diferenca": [round(r[3], 2) for r in soma_fonte],
        }
        st.dataframe(df_fonte, height=280, use_container_width=True)
    with col_total:
        st.subheader("Soma total")
        st.caption("Total geral por arquivo.")
        st.metric("Total CTB", f"{total_ctb:,.2f}")
        st.metric("Total Posicao", f"{total_posicao:,.2f}")
        st.metric("Diferenca (CTB - Posicao)", f"{diferenca_total:,.2f}")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Comparacao"
    ws1.append(["cod_reduzido", "nucleo_fonte", "valor_ctb", "valor_posicao", "diferenca"])
    for cod, nucleo, v_ctb, v_pos, diff in linhas:
        ws1.append([cod, nucleo, round(v_ctb, 2), round(v_pos, 2), round(diff, 2)])
    ws2 = wb.create_sheet("Soma por fonte")
    ws2.append(["nucleo_fonte", "total_ctb", "total_posicao", "diferenca"])
    for nucleo, v_ctb, v_pos, diff in soma_fonte:
        ws2.append([nucleo, round(v_ctb, 2), round(v_pos, 2), round(diff, 2)])
    ws3 = wb.create_sheet("Soma total")
    ws3.append(["total_ctb", "total_posicao", "diferenca"])
    ws3.append([round(total_ctb, 2), round(total_posicao, 2), round(diferenca_total, 2)])
    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    xlsx_buf.seek(0)
    st.download_button(
        "Download resultado (Excel - 3 abas)",
        xlsx_buf.getvalue(),
        file_name="comparacao_ctb_posicao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="xlsx",
    )





if __name__ == "__main__":
    main()
