from openpyxl import load_workbook
import os
import shutil
import tempfile
import zipfile
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from datetime import datetime

# ========== CONFIGURAÇÃO DO ARQUIVO DE DADOS (ORIGEM) ==========
# Altere aqui se o arquivo Excel de entrada mudar de estrutura.

# Aba/planilha de onde os dados são lidos (0 = primeira aba, ou nome exato da aba)
SOURCE_SHEET = 0

# Colunas obrigatórias no arquivo de dados (nomes exatos como no Excel)
REQUIRED_COLUMNS = [
    "credor",
    "cpf_cnpj",
    "valor_bruto",
    "valor_des",
    "Código de receita",
    "data",
]

# Mês por extenso (índice 1 = Janeiro, 12 = Dezembro)
MESES_EXTENSO = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]


def format_currency_br(value):
    """
    Formata valor para padrão brasileiro: ponto para milhar, vírgula para centavos, 2 decimais.
    Ex: 81634.0 -> "81.634,00" ; 3918.43 -> "3.918,43"
    Se o valor já vier como string no formato correto do Excel, preserva.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "0,00"
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return "0,00"
        # Já está no formato BR (tem vírgula e só dígitos/pontos)?
        if "," in s and s.replace(".", "").replace(",", "").replace(" ", "").isdigit():
            # Garante 2 casas decimais após a vírgula
            if "," in s:
                parte_int, _, parte_dec = s.partition(",")
                parte_dec = (parte_dec + "00")[:2]
                return f"{parte_int},{parte_dec}"
            return s
        try:
            value = float(s.replace(".", "").replace(",", "."))
        except (ValueError, AttributeError):
            return s
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    num = round(num, 2)
    int_part = int(num)
    dec_part = int(round((num - int_part) * 100)) % 100
    dec_str = f"{dec_part:02d}"
    sign = "-" if int_part < 0 else ""
    int_str = str(abs(int_part))
    if len(int_str) <= 3:
        return f"{sign}{int_str},{dec_str}"
    parts = []
    while int_str:
        parts.append(int_str[-3:])
        int_str = int_str[:-3]
    return f"{sign}{'.'.join(reversed(parts))},{dec_str}"


def sanitize_filename(name):
    """Remove caracteres inválidos do nome do arquivo (ex: / em 'ALGAR TELECOM S/A')."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return "credor"
    s = str(name).strip()
    for char in r'/\:*?"<>|':
        s = s.replace(char, "-")
    return s.strip("-") or "credor"


def parse_data_col(value):
    """Converte valor da coluna 'data' para datetime (aceita dd/mm/yyyy ou datetime do Excel)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    s = str(value).strip().split("\n")[0].strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d/%m/%Y")
    except ValueError:
        try:
            return pd.to_datetime(s)
        except Exception:
            return None


def month_extenso(dt):
    """Retorna o mês por extenso (ex: Maio)."""
    if dt is None:
        return ""
    if hasattr(dt, "month"):
        m = dt.month
    else:
        m = getattr(dt, "month", None) or 0
    return MESES_EXTENSO[m] if 1 <= m <= 12 else ""


def date_to_dd_mm_yyyy(dt):
    """Formata data para dd/mm/yyyy no formulário."""
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y")
    d = getattr(dt, "day", 1)
    m = getattr(dt, "month", 1)
    y = getattr(dt, "year", 1900)
    return f"{d:02d}/{m:02d}/{y:04d}"


def save_uploaded_file(uploaded_file):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            shutil.copyfileobj(uploaded_file, tmp_file)
            return tmp_file.name
    except Exception as e:
        st.error(f"Erro ao salvar o arquivo carregado: {e}")
        return None


def get_source_file_structure(uploaded_file):

    try:
        uploaded_file.seek(0)
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=SOURCE_SHEET, nrows=0)
        columns_found = list(df.columns)
        return sheet_names, SOURCE_SHEET, columns_found
    except Exception as e:
        return None, None, str(e)


def validate_source_columns(columns_found):
    """Verifica se todas as colunas obrigatórias existem no arquivo. Retorna (ok, listas de faltando/encontradas)."""
    missing = [c for c in REQUIRED_COLUMNS if c not in columns_found]
    return len(missing) == 0, missing, columns_found

def set_value(ws, address, value):
   
    cell = ws[address]
    for range_ in ws.merged_cells.ranges:
        if address in range_:
            
            ws.unmerge_cells(str(range_))
            
            top_left_cell_address = get_column_letter(range_.min_col) + str(range_.min_row)
            top_left_cell = ws[top_left_cell_address]
            top_left_cell.value = value
            ws.merge_cells(str(range_))
            return
    
    cell.value = value
  

def fill_and_save_form(group, name, payment_month, payment_date, template_file_path, cpf_cnpj=None):
    try:
        safe_name = sanitize_filename(name)
        safe_cnpj = sanitize_filename(str(cpf_cnpj) if cpf_cnpj is not None else "") or "cnpj"
        output_file_name = f'Preenchido_{safe_name}_{safe_cnpj}_Anexo_V - Comprovante de Retenção.xlsx'
        output_file_path = os.path.join(tempfile.gettempdir(), output_file_name)
        shutil.copy(template_file_path, output_file_path)

        wb = load_workbook(output_file_path)
        ws = wb.active  

        set_value(ws, 'A9', group['cpf_cnpj'].iloc[0])
        set_value(ws, 'D9', name)
        set_value(ws, 'D40', payment_date)
        set_value(ws, 'A40', "Hemerson Fernandes Soares")

        # Coluna auxiliar: data parseada para agrupar por mês/ano
        group = group.copy()
        group["_dt"] = group["data"].apply(parse_data_col)
        group = group[group["_dt"].notna()].copy()
        if group.empty:
            wb.save(output_file_path)
            wb.close()
            return output_file_path

        group["_ano"] = group["_dt"].apply(lambda x: getattr(x, "year", 1900))
        group["_mes"] = group["_dt"].apply(lambda x: getattr(x, "month", 1))

        # Soma total (para informações complementares em A35)
        total_bruto = group["valor_bruto"].sum()
        total_des = group["valor_des"].sum()
        info_complementares = f"Total valor pago (bruto): {format_currency_br(total_bruto)} | Total retido: {format_currency_br(total_des)}"
        set_value(ws, "A35", info_complementares)

        # Tabela: valores separados por mês (e por código de receita)
        group_por_mes = (
            group[["_ano", "_mes", "Código de receita", "valor_bruto", "valor_des"]]
            .groupby(["_ano", "_mes", "Código de receita"])
            .sum()
            .reset_index()
        )
        group_por_mes = group_por_mes.sort_values(["_ano", "_mes", "Código de receita"])

        start_row = 12
        for i, (_, row) in enumerate(group_por_mes.iterrows()):
            row_number = start_row + i
            mes_ref = MESES_EXTENSO[int(row["_mes"])] if 1 <= row["_mes"] <= 12 else ""
            set_value(ws, f"A{row_number}", mes_ref)
            set_value(ws, f"C{row_number}", row["Código de receita"])
            set_value(ws, f"E{row_number}", format_currency_br(row["valor_bruto"]))
            set_value(ws, f"G{row_number}", format_currency_br(row["valor_des"]))

        wb.save(output_file_path)
        wb.close()
        return output_file_path
    except Exception as e:
        st.error(f"Failed to process group {name}. Error: {e}")
        return None




def main():
    st.title("Processador de Formulários de Pagamento/IR RFB")

    uploaded_file = st.file_uploader("Escolha um arquivo Excel com os dados", type=["xls", "xlsx"])
    template_file = st.file_uploader("Escolha o arquivo de modelo do formulário", type="xlsx")
    payment_date = st.text_input("Insira a data do pagamento (dd/mm/aaaa):", "")
    file_name = st.text_input("Nome do arquivo para download:", "Preenchido_Formulario")

    # Mostra aba e colunas do arquivo de dados assim que um arquivo é carregado
    if uploaded_file:
        with st.expander("Estrutura do arquivo de dados (verificar se houve alteração)"):
            sheet_names, used_sheet, columns_found = get_source_file_structure(uploaded_file)
            if isinstance(columns_found, str):
                st.error(f"Erro ao ler o arquivo: {columns_found}")
            else:
                st.write("**Abas no arquivo:**", sheet_names)
                sheet_label = sheet_names[used_sheet] if isinstance(used_sheet, int) and used_sheet < len(sheet_names) else used_sheet
                st.write("**Aba utilizada (configuração atual):**", f"`SOURCE_SHEET = {repr(SOURCE_SHEET)}` → ", sheet_label)
                st.write("**Colunas esperadas (código):**", REQUIRED_COLUMNS)
                st.write("**Colunas encontradas na aba:**", columns_found)
                ok, missing, _ = validate_source_columns(columns_found)
                if ok:
                    st.success("Todas as colunas necessárias foram encontradas.")
                else:
                    st.error(f"Colunas faltando no arquivo: **{missing}**. Ajuste o arquivo ou a lista `REQUIRED_COLUMNS` no código.")

    process_button = st.button("Processar")

    if process_button:
        if not uploaded_file:
            st.warning("Envie o arquivo Excel com os dados.")
        elif not template_file:
            st.warning("Envie o arquivo de modelo do formulário.")
        elif not payment_date or not payment_date.strip():
            st.warning("Preencha a data do pagamento (dd/mm/aaaa).")
        else:
            template_file_path = save_uploaded_file(template_file)
            if template_file_path:
                uploaded_file.seek(0)
                sheet_names, used_sheet, columns_found = get_source_file_structure(uploaded_file)
                if isinstance(columns_found, str):
                    st.error(f"Não foi possível ler o arquivo de dados: {columns_found}")
                else:
                    ok, missing, _ = validate_source_columns(columns_found)
                    if not ok:
                        st.error(f"Arquivo de dados incompatível. Colunas faltando: {missing}. Verifique a aba e as colunas na seção acima.")
                    else:
                        uploaded_file.seek(0)
                        source_df = pd.read_excel(uploaded_file, sheet_name=SOURCE_SHEET)
                        filtered_source_df = source_df[REQUIRED_COLUMNS].copy()
                        grouped = filtered_source_df.groupby("cpf_cnpj")

                        output_files = []
                        for i, (cpf_cnpj, group) in enumerate(grouped):
                            group = group.reset_index(drop=True)
                            group_dates = group["data"].apply(parse_data_col)
                            valid = group_dates.notna()
                            if valid.any():
                                idx_last = group_dates[valid].idxmax()
                                row_last = group.loc[idx_last]
                                name = row_last["credor"]
                                dt = group_dates[idx_last]
                                payment_month = month_extenso(dt)
                            else:
                                name = group["credor"].iloc[0]
                                payment_month = ""

                            st.write(f"Processando grupo {i+1}/{len(grouped)}: {name} (CPF/CNPJ: {cpf_cnpj})")
                            output_file = fill_and_save_form(group, name, payment_month, payment_date.strip(), template_file_path, cpf_cnpj=cpf_cnpj)
                            if output_file:
                                output_files.append(output_file)

                        if output_files:
                            zip_file_path = os.path.join(tempfile.gettempdir(), f"{file_name}.zip")
                            with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                                for file in output_files:
                                    if os.path.exists(file):
                                        zipf.write(file, os.path.basename(file))
                                        try:
                                            os.remove(file)
                                        except OSError:
                                            pass

                            with open(zip_file_path, 'rb') as f:
                                st.download_button("Baixar Arquivos", f, file_name=f"{file_name}.zip")
                        else:
                            st.warning("Nenhum arquivo foi gerado. Verifique se o arquivo de dados tem as colunas preenchidas (cpf_cnpj, data, etc.) e se o modelo está correto.")

if __name__ == "__main__":
    main()
