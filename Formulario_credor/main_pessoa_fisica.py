from openpyxl import load_workbook
import os
import shutil
import tempfile
import zipfile
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

# ========== CONFIGURAÇÃO DO ARQUIVO DE DADOS (ORIGEM) - PESSOA FÍSICA ==========
# Altere aqui se o arquivo Excel de entrada mudar de estrutura.

# Aba/planilha de onde os dados são lidos (0 = primeira aba, ou nome exato da aba)
SOURCE_SHEET = 0

# Colunas obrigatórias no arquivo de dados (nomes exatos como no Excel)
REQUIRED_COLUMNS = [
    "credor",
    "cpf_cnpj",
    "valor_bruto",
    "valor_des",
    "Natureza do rendimento",
    "DESCRIÇÃO",
]


def format_currency_br(value):
    """
    Formata valor para padrão brasileiro: ponto para milhar, vírgula para centavos, 2 decimais.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "0,00"
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return "0,00"
        if "," in s and s.replace(".", "").replace(",", "").replace(" ", "").isdigit():
            if "," in s:
                parte_int, _, parte_dec = s.partition(",")
                parte_dec = (parte_dec + "00")[:2]
                return f"{parte_int},{parte_dec}"
            return s
        try:
            value = float(s.replace(".", "").replace(",", "."))
        except (ValueError, AttributeError):
            return s
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    num = round(num, 2)
    int_part = int(num)
    dec_part = int(round((num - int_part) * 100)) % 100
    dec_str = f"{dec_part:02d}"
    sign = "-" if int_part < 0 else ""
    int_str = str(abs(int_part))
    if len(int_str) <= 3:
        return f"{sign}{int_str},{dec_str}"
    parts = []
    while int_str:
        parts.append(int_str[-3:])
        int_str = int_str[:-3]
    return f"{sign}{'.'.join(reversed(parts))},{dec_str}"


def sanitize_filename(name):
    """Remove caracteres inválidos do nome do arquivo (ex: / em 'ALGAR TELECOM S/A')."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return "credor"
    s = str(name).strip()
    for char in r'/\:*?"<>|':
        s = s.replace(char, "-")
    return s.strip("-") or "credor"


def save_uploaded_file(uploaded_file):
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            shutil.copyfileobj(uploaded_file, tmp_file)
            return tmp_file.name
    except Exception as e:
        st.error(f"Erro ao salvar o arquivo carregado: {e}")
        return None


def get_source_file_structure(uploaded_file):
    try:
        uploaded_file.seek(0)
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=SOURCE_SHEET, nrows=0)
        columns_found = list(df.columns)
        return sheet_names, SOURCE_SHEET, columns_found
    except Exception as e:
        return None, None, str(e)


def validate_source_columns(columns_found):
    """Verifica se todas as colunas obrigatórias existem no arquivo."""
    # Aceita colunas com espaços: compara também versão "strip" do nome
    columns_stripped = {c.strip(): c for c in columns_found}
    missing = []
    for req in REQUIRED_COLUMNS:
        if req not in columns_found and req.strip() not in columns_stripped:
            missing.append(req)
    return len(missing) == 0, missing, columns_found


def _select_required_columns(df):
    """Seleciona colunas do dataframe que batem com REQUIRED_COLUMNS (por nome ou strip). Retorna df com colunas renomeadas para REQUIRED_COLUMNS."""
    rename = {}
    for req in REQUIRED_COLUMNS:
        for c in df.columns:
            if c.strip() == req.strip():
                rename[c] = req
                break
    if len(rename) != len(REQUIRED_COLUMNS):
        return None
    return df[[c for c in rename]].rename(columns=rename)


def set_value(ws, address, value):
    cell = ws[address]
    for range_ in ws.merged_cells.ranges:
        if address in range_:
            ws.unmerge_cells(str(range_))
            top_left_cell_address = get_column_letter(range_.min_col) + str(range_.min_row)
            top_left_cell = ws[top_left_cell_address]
            top_left_cell.value = value
            ws.merge_cells(str(range_))
            return
    cell.value = value


def fill_and_save_form_pf(group, name, payment_date, template_file_path):
    """
    Preenche o formulário de Pessoa Física para um credor.
    Mapeamento: credor->D11, cpf_cnpj->A11, valor_bruto->G15, valor_des->G20,
                DESCRIÇÃO->A13, "Hemerson Fernandes Soares"->A57, data->D57.
    """
    try:
        safe_name = sanitize_filename(name)
        output_file_name = f'Preenchido_{safe_name}_Pessoa_Fisica.xlsx'
        output_file_path = os.path.join(tempfile.gettempdir(), output_file_name)
        shutil.copy(template_file_path, output_file_path)

        wb = load_workbook(output_file_path)
        ws = wb.active

        # Por credor: somamos valor_bruto e valor_des; demais campos da primeira linha
        valor_bruto_total = group['valor_bruto'].sum()
        valor_des_total = group['valor_des'].sum()
        cpf_cnpj = group['cpf_cnpj'].iloc[0]
        descricao = group['DESCRIÇÃO'].iloc[0]
        if pd.isna(descricao):
            descricao = ""

        set_value(ws, 'D11', name)                    # credor
        set_value(ws, 'A11', cpf_cnpj)               # cpf_cnpj
        set_value(ws, 'G15', format_currency_br(valor_bruto_total))   # valor_bruto
        set_value(ws, 'G20', format_currency_br(valor_des_total))    # valor_des
        set_value(ws, 'A13', descricao)               # DESCRIÇÃO (par com Natureza do rendimento)
        set_value(ws, 'A57', "Hemerson Fernandes Soares")
        set_value(ws, 'D57', payment_date)

        wb.save(output_file_path)
        wb.close()
        return output_file_path
    except Exception as e:
        st.error(f"Erro ao processar o grupo {name}. Erro: {e}")
        return None


def main():
    st.title("Processador de Formulários de Pagamento - Pessoa Física")

    uploaded_file = st.file_uploader("Escolha um arquivo Excel com os dados", type=["xls", "xlsx"])
    template_file = st.file_uploader("Escolha o arquivo de modelo do formulário", type="xlsx")
    payment_date = st.text_input("Insira a data do pagamento (dd/mm/aaaa):", "")
    file_name = st.text_input("Nome do arquivo para download:", "Preenchido_Formulario_PF")

    if uploaded_file:
        with st.expander("Estrutura do arquivo de dados (verificar se houve alteração)"):
            sheet_names, used_sheet, columns_found = get_source_file_structure(uploaded_file)
            if isinstance(columns_found, str):
                st.error(f"Erro ao ler o arquivo: {columns_found}")
            else:
                st.write("**Abas no arquivo:**", sheet_names)
                sheet_label = sheet_names[used_sheet] if isinstance(used_sheet, int) and used_sheet < len(sheet_names) else used_sheet
                st.write("**Aba utilizada (configuração atual):**", f"`SOURCE_SHEET = {repr(SOURCE_SHEET)}` → ", sheet_label)
                st.write("**Colunas esperadas (código):**", REQUIRED_COLUMNS)
                st.write("**Colunas encontradas na aba:**", columns_found)
                ok, missing, _ = validate_source_columns(columns_found)
                if ok:
                    st.success("Todas as colunas necessárias foram encontradas.")
                else:
                    st.error(f"Colunas faltando no arquivo: **{missing}**. Ajuste o arquivo ou a lista `REQUIRED_COLUMNS` no código.")

    process_button = st.button("Processar")

    if process_button:
        if not uploaded_file:
            st.warning("Envie o arquivo Excel com os dados.")
        elif not template_file:
            st.warning("Envie o arquivo de modelo do formulário.")
        elif not payment_date or not payment_date.strip():
            st.warning("Preencha a data do pagamento (dd/mm/aaaa).")
        else:
            template_file_path = save_uploaded_file(template_file)
            if template_file_path:
                uploaded_file.seek(0)
                sheet_names, used_sheet, columns_found = get_source_file_structure(uploaded_file)
                if isinstance(columns_found, str):
                    st.error(f"Não foi possível ler o arquivo de dados: {columns_found}")
                else:
                    ok, missing, _ = validate_source_columns(columns_found)
                    if not ok:
                        st.error(f"Arquivo de dados incompatível. Colunas faltando: {missing}. Verifique a aba e as colunas na seção acima.")
                    else:
                        uploaded_file.seek(0)
                        source_df = pd.read_excel(uploaded_file, sheet_name=SOURCE_SHEET)
                        filtered_source_df = _select_required_columns(source_df)
                        if filtered_source_df is None:
                            st.error("Nem todas as colunas necessárias foram encontradas no arquivo. Verifique os nomes (com ou sem espaços).")
                        else:
                            grouped = filtered_source_df.groupby("credor")

                            output_files = []
                            for i, (name, group) in enumerate(grouped):
                                st.write(f"Processando grupo {i+1}/{len(grouped)}: {name}")
                                group = group.reset_index(drop=True)
                                output_file = fill_and_save_form_pf(group, name, payment_date.strip(), template_file_path)
                                if output_file:
                                    output_files.append(output_file)

                            if output_files:
                                zip_file_path = os.path.join(tempfile.gettempdir(), f"{file_name}.zip")
                                with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                                    for file in output_files:
                                        zipf.write(file, os.path.basename(file))
                                        os.remove(file)
                                with open(zip_file_path, 'rb') as f:
                                    st.download_button("Baixar Arquivos", f, file_name=f"{file_name}.zip")
                            else:
                                st.warning("Nenhum arquivo foi gerado. Verifique se o arquivo de dados tem linhas com a coluna 'credor' preenchida e se o modelo está correto.")


if __name__ == "__main__":
    main()
